import openpyxl

book = openpyxl.load_workbook("C:/Users/mahesmahesh/PycharmProjects/PythonSelfFramework/testData/TestData.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

# To write data to excel
# sheet.cell(row=2, column=2).value = "Mahesh"
# print(sheet.cell(row=2, column=2).value)

# To get the maximum row in a sheet
print(sheet.max_row)

# To get the maximum column in a sheet
print(sheet.max_column)

# Another way of fetching value from sheet
print(sheet['A6'].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "t3":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i, column=j).value
print(Dict)
