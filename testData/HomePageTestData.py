import openpyxl


class HomePageTestData:
    test_HomepageData = [{"firstname": "Mahesh", "email": "mahesh@test.com", "gender": "Female"},
                         {"firstname": "Suraj", "email": "suraj@test.com", "gender": "Male"},
                         {"firstname": "MaheshSur", "email": "maheshsu@test.com", "gender": "Female"}]

    @staticmethod
    def getTestData(testcase_name):
        Dict = {}
        book = openpyxl.load_workbook("C:/Users/mahesmahesh/PycharmProjects/PythonSelfFramework/testData/TestData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
